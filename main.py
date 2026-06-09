import json
import random
import sqlite3
from datetime import datetime
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ====== STOCK ======
FILE = "stock.json"

# ===== LOG FILE FOR STOCK =======
LOG_FILE = "data.json"


def load_data():
    try:
        with open(FILE, "r") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
        else: 
            return {}
    except:
        return {}
    

def save_data(data):
    with open(FILE, "w") as f:
        json.dump(data, f, indent = 4)


def load_logs():
    try:
        with open(LOG_FILE, "r") as f:
            return json.load(f)
    except:
        return []
    
    
def save_logs(logs):
    with open(LOG_FILE, "w") as f:
        json.dump(logs, f, indent= 4)


def save_transactions():
    with open("transactions.json", "w") as file:
        json.dump(transactions, file, indent = 4)

def load_transactions():
    global transactions
    try:
        with open("transactions.json", "r") as file:
            transactions = json.load(file)
    except:
        transactions = []


def clear_log():
    with open(LOG_FILE, "w"):
        listbox2.delete(0, END)
        pass
    

def show_logs():
    listbox2.delete(0, END)
    for x in logs:
        listbox2.insert(END, x)


def clear_stock():
    confirm = messagebox.askyesno(
        "confirm",
        "Delete all stock Records"
    )
    if confirm:
        stock.clear()
        save_data(stock)
        refresh_list()
        messagebox.showinfo("Success", "All stock deleted")


# ===== FUNCTION FOR ADDING ITEMS TO STOCK =====
def add_item():
    item = item_entry.get().strip().lower()
    qty = quantity_entry.get()
    
    if item == "" and qty == "":
        messagebox.showerror("input all fields")
    stock[item] = stock.get(item, 0) + int(qty)
    save_data(stock)
    messagebox.showinfo(f"Succesful, {item} added")
    refresh_list()

    data ={
        "Action" : "Added",
        "Item" : item,
        "qty" : qty,
        "Date" : str(datetime.now()),
    }
    logs.append(data)
    save_logs(logs)
    item_entry.delete(0, END)
    quantity_entry.delete(0, END)


# ==== FUNCTION FOR EDITING ITEM QUANTITY IN STOCK ====
def edit_item():
    item = item_entry.get().strip.lower()
    qty = quantity_entry.get()

    if item == "" or qty == "":
        messagebox.showerror("Enter item and quantity")
    elif item not in stock:
        messagebox.showerror("item not in stock")
        return
    try:
        qty = int(qty)
    except:
        messagebox.showerror("Quantity must be a number")
        return
    
    if item in stock:
        stock[item] = qty
        save_data(stock)
        messagebox.showinfo("Success",  f'{item} updated to {qty}')
        refresh_list()
    else:  
        messagebox.showerror("error", "item not found")
    (item_entry and quantity_entry).delete(0, END)

# ==== FUNCTION TO DELETE ITEM FROM STOCK ====
def delete_item():
    item = item_entry.get()

    if item in stock:
       del stock[item]
       save_data(stock)
       messagebox.showinfo(f"{item} Removed")
       refresh_list()
    else: 
        messagebox.showerror(f"{item} not in stock")
    
    data = {
        "Action" : "Deleted",
        "Item" : item,
        "Date" : str(datetime.now())
    }
    logs.append(data)
    save_logs(logs)


# ==== REFRESHING LISTS ======
def refresh_list():
    listbox1.delete(0, END)
    for item, qty in stock.items():
        #item = item.capitalize()
        listbox1.insert(END, f"{item} : {qty}")

# ==== REFRESHING ITEMS IN CART       
def refresh_cart():
    listbox2.delete(0, END)
    for data in cart:
        listbox2.insert(END, f"{data["item"]} - {data["qty"]}")


stock = load_data()
#print("xxxxxxxxxxxxxxxxxx")
print(type(stock))
for keys, values in stock.items():
    print(f"{keys} : {values}")
print("*****************************************")
for key in stock.keys():
    print(repr(key))
logs = load_logs()

#print("=============")
#print(type(logs))
#print(logs)

# ===== STOCK REPORTS TO EXCEL WORKSHEET =====
def export_to_excel():
    wb = Workbook()
# ------- STOCK SHEET -----
    ws = wb.active
    ws.title = "Stock Records"
    bold = Font(bold= True)

    ws.append(["Items, Quantity"])

    for item, qty in stock.items():
      ws.append([item, qty])
      ws['A' + str(ws.max_row)].font = bold
      ws['B' + str(ws.max_row)].font = bold

    ws.append([])
    ws.append([])

    for col in ws.columns:
        max_length= 0
        col_letter = col[0].column_letter
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))  
        
        ws.column_dimensions[col_letter].width = max_length + 4
    try:
        wb.save("inventory_report.xlsx")
        messagebox.showinfo("Success", "Stock Records Exported to Excel")
    except PermissionError:
        messagebox.showwarning("Close the Excel file first")

 
# ==== EXPORT RENTALS TO EXCEL WORKSHEET =====
def export_rentals():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentals Record"

    bold = Font(bold = True)
 
    for trans in transactions:
       ws.append([f"Customer: {trans['customer']}"])
       ws['A' + str(ws.max_row)].font = bold

       ws.append([f"Phone: {trans['phone']}"])
       ws.append([f"Rent ID: {trans['rent_id']}"])
       ws.append([f"Date: {trans['date']}"])

       ws.append([f"Item", "Quantity"])

       ws["A" + str(ws.max_row)].font = bold
       ws["B" + str(ws.max_row)].font = bold

       for item in trans["items"]:
           ws.append([
               item["item"],
               item["qty"]
           ])
       ws.append([])
       ws.append([])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15

    wb.save("Rentals.xlsx")     


    


# ====  FUNCTION FOR CART
cart = []

def add_to_cart():
    item = item_entry.get().strip().lower()
    print(item)
    qty = quantity_entry.get()

    if item == '' or qty == '':
        messagebox.showerror("all fields are required")
        return
    elif item not in stock:
        messagebox.showerror(f"{item} not found")
        return
    elif item in stock and stock[item] < int(qty):
        messagebox.showerror("Not Enough Stock")
        return

    cart.append({
        "item" : item,
        "qty" : qty
    })
    listbox2.insert(END, f"{item} : {qty}")
    item_entry.delete(0, END)
    quantity_entry.delete(0, END)

# ==== FUNCTION TO DELETE FROM CART =====
def del_from_cart():  
    item = item_entry.get().strip().lower()
    qty = quantity_entry.get()
    
    found = False
    for data in cart:
        if data["item"] == item:
            cart.remove(data)
            messagebox.showinfo(f"{item} removed from cart")
            found = True
            refresh_cart()
            break
    if not found:
        messagebox.showerror(f"{item} not found")
    # print(cart)
    item_entry.delete(0, END)
    quantity_entry.delete(0, END)


# ==== RENTALS HISTORY =====
transactions = []

load_transactions()

def delete_transactions():
    for trans in transactions:
        transactions.clear()
        save_transactions()

# ===== FUNCTION TO ISSUE OUT ITEMS ======
def issue():
    customer = customer_entry.get().strip().lower()
    phone = contact_entry.get()
    #event = event_entry.get()

    if not cart:
        messagebox.showerror("Error", "Cart is Empty")
        return
    elif customer == "" or phone == "":
        messagebox.showerror("Customer Details is required")
        return
     # creating an id for returning items
    num = "0123456789"
    low = 'abcdefghijklmnopqrstuvwxyz'
    high = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    chars = num + low + high
    length = 5
    global rent_id
    rent_id = ''.join(random.sample(chars, length))
    #print(f"rent_id = {rent_id}")

    # checking and reducing stock
    for data in cart:
        item = data["item"]
        qty = data["qty"]

        if item in stock and stock[item] >= int(qty):
            stock[item] -= int(qty)
        else:
            messagebox.showerror("Error", f"Not enough {item}")
            return
    save_data(stock)

    transaction = {
        "customer" : customer,
        "phone" : phone,
        "rent_id": rent_id,
       # "event" : event,
        "date" : str(datetime.now().strftime("%d-%m-%Y  %I:%M %p")),
        "items" : cart.copy(),
        "returned" : False
    }
    transactions.append(transaction)
    save_transactions()
    # print(cart)

    # display
    items = "\n".join(f"{item['item']} - {item['qty']}" for item in cart)
    message = f"Confirm your items! as there will be no room for arguments afterwards"
    # print(items)
    receipt = (
     f"""
     customer = {customer}
     phone = {phone}
     rent_id = {rent_id}
     date = {datetime.now().strftime("%d-%m-%Y  %I:%M %p")}
     items = {items}
     NOTE: = {message}
     """
     )
    print(receipt)
    order.config(text = receipt)
    
    # clear cart
    cart.clear()
    listbox2.delete(0, END)
    messagebox.showinfo("Success", "items issued successfully")

    customer_entry.delete(0, END) 
    contact_entry.delete(0, END)
    #event_entry.delete(0, END)
    refresh_list()


# ====== FUNCTION TO RETURN ITEMS RENTED OUT =====
def return_item(phone, rent_id):
    
    for trans in transactions:
        # print(transactions)

        if  (trans['phone'] == phone and trans['rent_id'] == rent_id): 
          
           # CHECK IF ALREADY RETURNED
           if trans['returned']:
               messagebox.showinfo('Info', 'items already returned')
               return
           for item in trans['items'] :
               stock[item['item']] += int(item["qty"])

           # MARK AS RETURNED
           trans['returned'] = True
           save_transactions()
           # print(transactions)

           # SAVE UPDATED STOCK
           save_data(stock)
           messagebox.showinfo("Success", f'{trans["customer"]} returned items successfully')
           contact_entry.delete(0, END)
           ID_entry.delete(0, END)
           return
        
    messagebox.showerror("Error", "customer not found")
            
# print(f'this is transations == {transactions}')
        
# ========================


# ====== GUI WINDOW =======
base = Tk()
base.title("Store Inventory Record System")
base.geometry("500x500")

frm1 = Frame(base, bg = "teal")
frm1.pack(fill = BOTH, expand = True, ipadx= 10, ipady = 5, pady = 5, side = LEFT)

frm2 = Frame(frm1, bg = "blue")
frm2.pack(fill = BOTH, expand = True,padx = 10, pady = 10, side = LEFT)

label_frame = Frame(frm2, bg = "white")
label_frame.pack(fill= BOTH)

Label(label_frame, text = "Item Name").pack( padx= 3, pady = 5)
item_entry = Entry(label_frame)
item_entry.pack(padx = 3, pady = 3)

Label(label_frame, text = "Quantity").pack( pady = 5)
quantity_entry = Entry(label_frame)
quantity_entry.pack(padx = 3, pady = 3)

frame_button = Frame(frm2, bg = "green")
frame_button.pack(anchor = "center")

customer_frame = Frame(frm2, bg = "orange")
Label(customer_frame, text = "Customer Name").pack(padx= 10)
customer_entry = Entry(customer_frame)
customer_entry.pack(pady= 5)

Label(customer_frame, text = "Phone").pack(padx= 10)
contact_entry = Entry(customer_frame)
contact_entry.pack(pady = 5)

Label(customer_frame, text ="Rent ID").pack(padx= 10)
ID_entry = Entry(customer_frame)
ID_entry.pack(pady = 5)

customer_frame.pack(fill= BOTH, ipadx= 10, ipady = 5, pady = 5, padx = 10 )



# ====== BUTTONS ======
btn1 = Button(frame_button, text = 'Add Item', command = add_item).grid(row = 1, column = 1, padx= 10, pady = 10)
btn2 = Button(frame_button, text = 'Export Stock', command = export_to_excel).grid(row = 1, column = 2, padx= 10, pady = 10)
btn3 = Button(frame_button, text = 'Edit Item', command = edit_item).grid(row = 1, column = 3, padx = 10, pady = 10)
bnt4 = Button(frame_button, text = 'Delete Item', command = delete_item).grid(row = 1, column = 4, padx = 10, pady = 10)
bnt5 = Button(frame_button, text = 'Refresh List', command = refresh_list).grid(row = 2, column = 1, padx = 10, pady = 10)
bnt6 = Button(frame_button, text = 'Show Logs', command = show_logs).grid(row = 2, column = 2, padx = 10, pady = 10)
bnt7 = Button(frame_button, text = 'Clear Logs', command = clear_log).grid(row = 2, column = 3, padx = 10, pady = 10)
bnt8 = Button(frame_button, text = 'Export Rentals', command = export_rentals).grid(row = 2, column = 4, padx = 10, pady = 10)
bnt9 = Button(frame_button, text = 'Add to Cart', command = add_to_cart).grid(row = 3, column = 1, padx = 10, pady = 10)
bnt10 = Button(frame_button, text = 'Issue All', command = issue).grid(row = 3, column = 2, padx = 10, pady = 10)
btn11 = Button(frame_button, text = "Return item", command = lambda: return_item(contact_entry.get(), ID_entry.get())).grid(row = 3, column = 3, padx = 10, pady = 10)
btn12 = Button(frame_button, text = "Delete Cart", command = del_from_cart).grid(row = 3, column = 4, padx= 0, pady= 10)
# =====================================


frm3 = Frame(frm1, bg = "white")
frm3.pack(fill = BOTH, expand = True, ipadx = 10, ipady = 10, padx = 10, pady = 10, side = LEFT)

order = Label(frm2, bg = "white", text = "ORDER LIST", height= 30)
order.pack(fill = BOTH, expand = True,padx = 10, pady = 5, side = LEFT)

####### LISTBOXS
listbox1 = Listbox(frm3, bg = 'white', font = ("Arial, 20"), fg = "blue", height = 5)
listbox1.pack(fill = BOTH,ipadx = 20,  expand = True)

listbox2 = Listbox(frm3, bg = 'white', font = ("Arial, 15"), fg = "blue", height = 9)
listbox2.pack(fill = BOTH,ipadx = 20,  expand = True)


base.mainloop()